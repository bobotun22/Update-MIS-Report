import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog, scrolledtext
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd
import re

# --- Function Definitions ---
def sheet_exists(wb, name):
    return name in wb.sheetnames

def prepare_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
    return wb.create_sheet(sheet_name)

def find_column_index(header, keyword):
    for i, h in enumerate(header):
        if h and keyword.lower() in str(h).lower():
            return i
    return -1

def to_float_safe(value):
    if value is None:
        return 0.0
    try:
        s = str(value).replace(",", "").replace("MMK", "").strip()
        return float(s)
    except Exception:
        return 0.0

def add_terminal_details_sheet(wb_main, terminal_aggregation):
    """Add a sheet with terminal details using aggregated transaction data"""
    ws_terminal = prepare_sheet(wb_main, "Terminal Details")
    
    # Add headers with the new columns
    headers = ["S/N", "Terminal ID", "Location", "Total Transaction Count", "Total use card", "On us card", "Off us card", "Total Amount", "Fee", "SUM", "On us amount", "Off us amount"]
    ws_terminal.append(headers)
    
    # Terminal data
    terminal_info = [
        [1, "09950001", "Dagon (1)"],
        [2, "09950002", "Botahtaung"],
        [3, "09950003", "MIP"],
        [4, "09950004", "MDY"],
        [5, "09950005", "NPT MOC 40"],
        [6, "09950006", "Pathein"],
        [7, "09950007", "Naypyitaw Br"],
        [8, "09950008", "Mawlamyaing"],
        [9, "09950009", "Shwe Pauk Kan"],
        [10, "09950010", "Magway"],
        [11, "09950013", "Dagon (2)"],
        [12, "09950014", "Monywa"],
        [13, "09950015", "Hpa-an"],
        [14, "09950016", "Dawei"],
        [15, "09950017", "DSK"],
        [16, "09950018", "BGO"],
        [17, "09950019", "MKN"],
        [18, "09950021", "HTY"],
        [19, "09950024", "TGI"],
        [20, "09950025", "MICT Park"],
        [21, "09950026", "NPT MOC 11"],
        [22, "09950027", "NPT MOC 11 (2)"],
        [23, "09950028", "NPT MOC 40 (2)"],
        [24, "09950029", "NPT Hluttaw"]
    ]
    
    # Grouping lists for final row summaries
    ygn_branch_ids = ["09950001", "09950002", "09950009", "09950013", "09950017", "09950021"]
    ygn_public_ids = ["09950003", "09950025"]
    other_branch_ids = ["09950004", "09950006", "09950007", "09950008", "09950010", "09950014", "09950015", "09950016", "09950018", "09950019", "09950024"]
    other_public_ids = ["09950005", "09950026", "09950027", "09950028", "09950029"]

    # Initializing variables to track grand totals
    grand_totals = {
        'count': 0, 'use_card': 0, 'on_us_card': 0, 'off_us_card': 0,
        'amount': 0.0, 'fee': 0.0, 'sum': 0.0, 'on_us_amt': 0.0, 'off_us_amt': 0.0
    }

    # Initializing variables to track custom group sums (Amount SUM and Total Use Card)
    
    ygn_branch_sum = 0.0
    ygn_branch_cards = 0  
    
    ygn_public_sum = 0.0
    ygn_public_cards = 0
    
    other_branch_sum = 0.0
    other_branch_cards = 0    
    
    other_public_sum = 0.0
    other_public_cards = 0
    

    
    # Add terminal data to the sheet using aggregated data
    for info in terminal_info:
        sn, terminal_id, location = info
        
        # Get aggregated data for this terminal
        terminal_data = terminal_aggregation.get(terminal_id, {
            'total_transaction_count': 0,
            'total_transaction_amount': 0.0,
            'total_fee': 0.0,
            'on_us_count': 0,
            'off_us_count': 0,
            'on_us_amount': 0.0,
            'off_us_amount': 0.0,
            'unique_card_count': 0,
            'card_numbers': set()
        })
        
        transaction_count = terminal_data['total_transaction_count']
        total_amount = terminal_data['total_transaction_amount']
        total_fee = terminal_data['total_fee']
        sum_value = total_amount - total_fee
        
        on_us_count = terminal_data['on_us_count']
        off_us_count = terminal_data['off_us_count']
        total_use_card = terminal_data['unique_card_count']
        on_us_amount = terminal_data['on_us_amount']
        off_us_amount = terminal_data['off_us_amount']
        
        # Accumulate Grand Totals
        grand_totals['count'] += transaction_count
        grand_totals['use_card'] += total_use_card
        grand_totals['on_us_card'] += on_us_count
        grand_totals['off_us_card'] += off_us_count
        grand_totals['amount'] += total_amount
        grand_totals['fee'] += total_fee
        grand_totals['sum'] += sum_value
        grand_totals['on_us_amt'] += on_us_amount
        grand_totals['off_us_amt'] += off_us_amount

        # Accumulate sums and card counts for custom groups
        if terminal_id in ygn_public_ids:
            ygn_public_sum += sum_value
            ygn_public_cards += total_use_card
        elif terminal_id in ygn_branch_ids:
            ygn_branch_sum += sum_value
            ygn_branch_cards += total_use_card
        elif terminal_id in other_public_ids:
            other_public_sum += sum_value
            other_public_cards += total_use_card
        elif terminal_id in other_branch_ids:
            other_branch_sum += sum_value
            other_branch_cards += total_use_card
            
        ws_terminal.append([
            sn, terminal_id, location, transaction_count, total_use_card, on_us_count, off_us_count,
            total_amount, total_fee, sum_value, on_us_amount, off_us_amount
        ])
    
    # Append the Grand Total Row (Total SUM of all terminals)
    grand_total_row_idx = ws_terminal.max_row + 1
    ws_terminal.append([
        "", "", "Total SUM", 
        grand_totals['count'], grand_totals['use_card'], grand_totals['on_us_card'], grand_totals['off_us_card'],
        grand_totals['amount'], grand_totals['fee'], grand_totals['sum'], grand_totals['on_us_amt'], grand_totals['off_us_amt']
    ])

    # Add an empty row for visual separation
    ws_terminal.append([])
    
    # Append the specific category summary definitions (Now including Total Use Card in column 5)
    summary_start_row = ws_terminal.max_row + 1
    ws_terminal.append(["", "", "YGN Branch Total SUM", "", ygn_branch_cards, "", "", "", "", ygn_branch_sum, "", ""])
    ws_terminal.append(["", "", "YGN Public Total SUM", "", ygn_public_cards, "", "", "", "", ygn_public_sum, "", ""])
    ws_terminal.append(["", "", "Other Branch Total SUM", "", other_branch_cards, "", "", "", "", other_branch_sum, "", ""])
    ws_terminal.append(["", "", "Other Public Total SUM", "", other_public_cards, "", "", "", "", other_public_sum, "", ""])
    summary_end_row = ws_terminal.max_row

    # Format numbers and apply bold stylings across rows cleanly
    for row in range(2, ws_terminal.max_row + 1):
        # 1. Standard Data Rows Layout 
        if row < grand_total_row_idx:
            for col in [4, 5, 6, 7]:
                ws_terminal.cell(row=row, column=col).number_format = '#,##0'
            for col in [8, 9, 10, 11, 12]:
                ws_terminal.cell(row=row, column=col).number_format = '#,##0.00'
        
        # 2. Grand Total Row Highlight Styling
        elif row == grand_total_row_idx:
            for col in range(1, 13):
                cell = ws_terminal.cell(row=row, column=col)
                cell.font = Font(bold=True)
                if col in [4, 5, 6, 7]:
                    cell.number_format = '#,##0'
                elif col in [8, 9, 10, 11, 12]:
                    cell.number_format = '#,##0.00'
                    
        # 3. Location Category Summaries Block Styling
        elif summary_start_row <= row <= summary_end_row:
            ws_terminal.cell(row=row, column=3).font = Font(bold=True)
            
            # Format and Bold the specific card counter column (Col 5)
            card_cell = ws_terminal.cell(row=row, column=5)
            card_cell.number_format = '#,##0'
            card_cell.font = Font(bold=True)
            
            # Format and Bold the specific Category SUM column (Col 10)
            sum_cell = ws_terminal.cell(row=row, column=10)
            sum_cell.number_format = '#,##0.00'
            sum_cell.font = Font(bold=True)
            
    # Auto-adjust column widths safely
    for column in ws_terminal.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_terminal.column_dimensions[column_letter].width = adjusted_width

def add_terminal_transaction_sheets(wb_main, terminal_transactions, header):
    """Add a sheet for each terminal with its detailed transactions"""
    for terminal_id, transactions in terminal_transactions.items():
        if not transactions:  
            continue
            
        sheet_name = f"T_{terminal_id}"
        ws = prepare_sheet(wb_main, sheet_name)
        ws.append(header)
        
        for row in transactions:
            ws.append([cell.value for cell in row])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 30)  
            ws.column_dimensions[column_letter].width = adjusted_width

def process_transactions(ws_main, wb_main, header, amount_idx, fee_idx, card_idx, refno_idx, terminal_id_idx, status_idx, tranx_amount_idx, terminal_aggregation, terminal_transactions, terminal_ids, reversal_ref_nos):
    """Process transactions and aggregate data by terminal"""
    rows_to_delete = []
    
    for row_idx in range(2, ws_main.max_row + 1):  
        row = ws_main[row_idx]
        if status_idx != -1 and row[status_idx].value is not None:
            status_value = str(row[status_idx].value).strip()
            if status_value == "9":
                rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws_main.delete_rows(row_idx)
    
    for row_idx in range(2, ws_main.max_row + 1):  
        row = ws_main[row_idx]
        ref_no = str(row[refno_idx].value).strip() if row[refno_idx].value else ""
        
        if ref_no in reversal_ref_nos:
            continue

        terminal_id = str(row[terminal_id_idx].value).strip() if terminal_id_idx != -1 and row[terminal_id_idx].value else ""
        amount = to_float_safe(row[amount_idx].value)
        fee = to_float_safe(row[fee_idx].value)
        card_number = str(row[card_idx].value).strip() if card_idx != -1 and row[card_idx].value else ""
        
        if terminal_id in terminal_ids:
            terminal_transactions[terminal_id].append(row)
            terminal_aggregation[terminal_id]['total_transaction_amount'] += amount
            terminal_aggregation[terminal_id]['total_fee'] += fee
            terminal_aggregation[terminal_id]['total_transaction_count'] += 1
            
            if card_number and card_number.startswith("950316"):
                terminal_aggregation[terminal_id]['on_us_count'] += 1
                terminal_aggregation[terminal_id]['on_us_amount'] += amount
            else:
                terminal_aggregation[terminal_id]['off_us_count'] += 1
                terminal_aggregation[terminal_id]['off_us_amount'] += amount
            
            if card_number:
                terminal_aggregation[terminal_id]['card_numbers'].add(card_number)
                terminal_aggregation[terminal_id]['unique_card_count'] = len(terminal_aggregation[terminal_id]['card_numbers'])
        
        if status_idx != -1 and row[status_idx].value is not None:
            original_status = str(row[status_idx].value)
            cleaned_status = original_status.replace('9', '')
            row[status_idx].value = cleaned_status
        
        if tranx_amount_idx != -1 and row[tranx_amount_idx].value is not None:
            original_amount = str(row[tranx_amount_idx].value)
            cleaned_amount = original_amount.replace('MMK', '')
            row[tranx_amount_idx].value = cleaned_amount

def process_transaction_data(ws_main, wb_main):
    """Process transaction data and create sheets"""
    header = None
    header_row = 1
    
    for row_num in range(1, min(4, ws_main.max_row + 1)):
        test_header = [cell.value for cell in ws_main[row_num]]
        header_text = " ".join([str(h).lower() for h in test_header if h])
        if any(keyword in header_text for keyword in ["terminal", "amount", "fee", "ref"]):
            header = test_header
            header_row = row_num
            break
    
    if header is None:
        header = [cell.value for cell in ws_main[1]]
    
    amount_idx = find_column_index(header, "amount")
    fee_idx = find_column_index(header, "fee")
    card_idx = find_column_index(header, "card")
    refno_idx = find_column_index(header, "ref")
    
    terminal_id_idx = -1
    for keyword in ["terminal id", "terminal", "tid"]:
        terminal_id_idx = find_column_index(header, keyword)
        if terminal_id_idx != -1:
            break

    status_idx = find_column_index(header, "status")
    
    tranx_amount_idx = find_column_index(header, "tranx amount")
    if tranx_amount_idx == -1:
        tranx_amount_idx = find_column_index(header, "transaction amount")
    if tranx_amount_idx == -1:
        tranx_amount_idx = amount_idx  
    
    if amount_idx == -1 or fee_idx == -1 or refno_idx == -1:
        messagebox.showerror("Error", "❌ Could not find required columns.")
        return None, None

    if terminal_id_idx == -1:
        messagebox.showerror("Error", "❌ Could not find terminal column.")
        return None, None

    terminal_ids = [
        "09950001", "09950002", "09950003", "09950004", "09950005", "09950006", "09950007", "09950008", "09950009", "09950010",
        "09950013", "09950014", "09950015", "09950016", "09950017", "09950018", "09950019", "09950021", "09950024", "09950025",
        "09950026", "09950027", "09950028", "09950029"
    ]
    
    terminal_aggregation = {}
    terminal_transactions = {}
    for tid in terminal_ids:
        terminal_aggregation[tid] = {
            'total_transaction_count': 0,
            'total_transaction_amount': 0.0,
            'total_fee': 0.0,
            'on_us_count': 0,
            'off_us_count': 0,
            'on_us_amount': 0.0,
            'off_us_amount': 0.0,
            'unique_card_count': 0,
            'card_numbers': set()
        }
        terminal_transactions[tid] = []

    reversal_ref_nos = set()
    for row in ws_main.iter_rows(min_row=header_row + 1):
        txt = " ".join([str(cell.value).strip() if cell.value else '' for cell in row])
        ref_no = str(row[refno_idx].value).strip() if row[refno_idx].value else ""
        if "reversal" in txt.lower() and ref_no:
            reversal_ref_nos.add(ref_no)

    process_transactions(
        ws_main, wb_main, header, amount_idx, fee_idx, card_idx, refno_idx, 
        terminal_id_idx, status_idx, tranx_amount_idx, terminal_aggregation, terminal_transactions, terminal_ids, reversal_ref_nos
    )
    
    add_terminal_transaction_sheets(wb_main, terminal_transactions, header)
    return terminal_aggregation, terminal_ids

def convert_xls_to_xlsx(xls_path):
    try:
        xls = pd.ExcelFile(xls_path, engine='xlrd')
        first_sheet = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=first_sheet)
        xlsx_path = xls_path + "_converted.xlsx"
        df.to_excel(xlsx_path, sheet_name="Sheet1", index=False)
        return xlsx_path
    except Exception as e:
        messagebox.showerror("Conversion Error", f"❌ Failed to convert .xls to .xlsx\n\n{e}")
        return None

def load_data(progress=None, status_label=None):
    if progress:
        progress["value"] = 0
        progress.update()
        status_label.config(text="📂 Opening file...")

    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xls *.xlsx *.xlsm"), ("All Files", "*.*")]
    )

    if not file_path:
        messagebox.showwarning("No File", "No file selected.")
        if status_label:
            status_label.config(text="⚠️ Cancelled.")
        return

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xls":
        if status_label:
            status_label.config(text="🔄 Converting .xls to .xlsx...")
        converted = convert_xls_to_xlsx(file_path)
        if not converted:
            return
        file_path = converted

    try:
        if status_label:
            status_label.config(text="📊 Loading workbook...")
        wb_data = load_workbook(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"❌ Failed to open the Excel file.\n\n{e}")
        if status_label:
            status_label.config(text="❌ Error opening file.")
        return

    ws_data = wb_data["Sheet1"] if "Sheet1" in wb_data.sheetnames else wb_data.active

    wb_main = Workbook()
    ws_main = wb_main.active
    ws_main.title = "Main"

    for row in ws_data.iter_rows():
        ws_main.append([cell.value for cell in row])
    wb_data.close()

    if status_label:
        status_label.config(text="⚙️ Processing...")
    if progress:
        progress["value"] = 50
        progress.update()

    result = process_transaction_data(ws_main, wb_main)
    if result is None or result[0] is None:
        return
        
    terminal_aggregation, terminal_ids = result

    if progress:
        progress["value"] = 90
        progress.update()

    add_terminal_details_sheet(wb_main, terminal_aggregation)

    if progress:
        progress["value"] = 100
        progress.update()

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        title="Save Processed File As"
    )

    if not save_path:
        messagebox.showinfo("Canceled", "Save operation canceled.")
        if status_label:
            status_label.config(text="⚠️ Save canceled.")
        return

    try:
        wb_main.save(save_path)
        if status_label:
            status_label.config(text="✅ Saved successfully.")
        messagebox.showinfo("Success", f"✅ File saved to:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"❌ Failed to save the file.\n\n{e}")
        if status_label:
            status_label.config(text="❌ Error saving file.")

# --- GUI Setup ---
root = tk.Tk()
root.title("💳 Weekly ATM Transaction MIS Report")
root.geometry("520x380")
root.resizable(False, False)
root.configure(bg="#f5f6fa")

frame = tk.Frame(root, padx=30, pady=20, bg="#f5f6fa")
frame.pack(expand=True, fill="both")

logo_label = tk.Label(frame, text="Excel Transaction Processor", font=("Arial", 17, "bold"), fg="#2f3640", bg="#f5f6fa")
logo_label.pack(pady=(0, 12))

desc_label = tk.Label(frame, text="Weekly & Monthly ATM Transaction Monitoring Report", font=("Arial", 11), fg="#353b48", bg="#f5f6fa")
desc_label.pack(pady=(0, 10))

progress = ttk.Progressbar(frame, orient="horizontal", length=400, mode="determinate")
progress.pack(pady=(10, 5))

status_label = tk.Label(frame, text="Ready.", font=("Arial", 12), fg="#40739e", bg="#f5f6fa")
status_label.pack(pady=(4, 14))

btn = tk.Button(
    frame,
    text="📥 Import Data Excel File (FeelSwitch)",
    command=lambda: load_data(progress, status_label),
    width=40, height=2, bg="#2410D9", fg="white", font=("Arial", 12, "bold"),
    activebackground="#324ebd", bd=0, relief="flat", highlightthickness=0
)
btn.pack(pady=(8, 8))

credit_label = tk.Label(frame, text="© Version 3 | Developed by Bo Bo Tun", font=("Arial", 9, "italic bold"), fg="#a5b1c2", bg="#f5f6fa")
credit_label.pack(side="bottom", pady=(18, 0)) 

root.mainloop()